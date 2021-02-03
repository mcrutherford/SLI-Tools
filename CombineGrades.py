"""

@file: CombineGrades.py
@author: Mark Rutherford [ mcr5801 ]
@created: 8/30/2018 4:35 PM
@description:
    Combine completed grading rubrics from all students into one master xlsx to add to mycourses.

    Validated for Python 3.8, but should work for Python 3.x. Requires openpyxl, which can be installed with the command
    'pip install openpyxl'

"""
import openpyxl  # pip install openpyxl
import openpyxl.utils  # pip install openpyxl
import tkinter as tk
from tkinter import filedialog
import os


def get_rubrics(input_folder_path):
    """
    Get all students' rubrics.

    :param input_folder_path: The main folder containing students' folders
    :return: A dictionary containing the student's name and their rubric
    """
    rubrics = {}
    for student_name in os.listdir(input_folder_path):  # loop through all the files and folders
        if os.path.isdir(
                os.path.join(input_folder_path, student_name)):  # check whether the current object is a folder or not
            if os.path.isfile(os.path.join(input_folder_path, student_name, 'Rubric.xlsx')):
                rubrics[student_name] = os.path.join(input_folder_path, student_name, 'Rubric.xlsx')
    return rubrics


def combine_rubrics(rubrics, outputws):
    """
    Combine individual rubrics into one master grading sheet.

    :param rubrics: The students' rubric files
    :param outputws: The worksheet to output to
    """
    for index, name in enumerate(rubrics.keys()):
        rubricpath = rubrics[name]
        rubricwb = openpyxl.load_workbook(rubricpath)
        rubricws = rubricwb.worksheets[0]
        split_name = name.split('_')
        name = split_name[0] + ', ' + split_name[1]

        # Iterate over each row to calculate total score and comment for mycourses
        comment_string = ''
        total_points_achieved = 0
        for row in range(2, rubricws.max_row+1):
            # Calculate points achieved
            points_deducted = int(rubricws.cell(row, 3).value) if rubricws.cell(row, 3).value else 0
            points_possible = int(rubricws.cell(row, 2).value) if rubricws.cell(row, 2).value else 0
            points_achieved = points_possible - points_deducted
            total_points_achieved += points_achieved

            # Add the comment to the full comment string
            comment_string += rubricws.cell(row, 1).value  # the requirement
            if row != rubricws.max_row:
                comment_string += ' [' + str(points_achieved) + '/' + str(points_possible) + ']'  # points lost
            if rubricws.cell(row, 4).value:
                comment = ':\n' + rubricws.cell(row, 4).value  # grader comment
                if comment[-1] != '\n':
                    comment += '\n'
                # comment.replace('\n', '\n\t')  # Tab in the comment
                comment_string += comment
            elif row < rubricws.max_row+1:
                comment_string += '\n'
        comment_string = comment_string.strip()

        # add line to the output worksheet
        outputws.cell(index+1, 1).value = name
        outputws.cell(index+1, 2).value = total_points_achieved
        outputws.cell(index+1, 3).value = comment_string


def main():
    root = tk.Tk()
    root.withdraw()

    # Prompt for the main folder containing student files
    input_folder_path = filedialog.askdirectory(title='Select folder containing student assignment folders')
    if input_folder_path == '':
        return

    outputwb = openpyxl.Workbook()
    outputws = outputwb['Sheet']

    outputws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 30
    outputws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 15
    outputws.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 60

    rubrics = get_rubrics(input_folder_path)

    combine_rubrics(rubrics, outputws)

    outputwb.save(os.path.join(input_folder_path, 'AllStudents.xlsx'))


if __name__ == "__main__":
    main()
    print("\ndone")
