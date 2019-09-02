"""

@file: parsexl.py
@author: Mark Rutherford [ mcr5801 ]
@created: 8/30/2018 4:35 PM
@description:
    

"""
import openpyxl # Not standard
import openpyxl.utils

import tkinter as tk
from tkinter import filedialog
import os


def create_individual_xl(rubric_data, location, easy_access_vars):
    # wb = openpyxl.Workbook(location + '/Rubric.xlsx')
    wb = openpyxl.Workbook()
    ws = wb['Sheet']

    # ws.set_column(0, 3, cell_format = wrap)
    ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 30
    ws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 14
    ws.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 10
    ws.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 70

    col_names = ['Requirement', 'Points available', 'Deductions', 'Comments']
    for i in range(len(col_names)):
        ws.cell(1, i+1).value = col_names[i]
    row = 1
    for i, key in enumerate(rubric_data['requirements'].keys()):
        row = i+2
        ws.cell(row, 1).value = key
        ws.cell(row, 2).value = rubric_data['requirements'][key]["value"]
        if ("auto deductions" in rubric_data['requirements'][key]):
            deduct_sum = 0
            comment = ""
            for autodeduct in rubric_data['requirements'][key]['auto deductions']:
                print(autodeduct['command'])
                if not eval(autodeduct['command']):
                    deduct_sum += autodeduct['penalty']
                    comment += ' -' + str(autodeduct['penalty']) + 'p: ' + autodeduct['desc'] + '\n'
            ws.cell(row, 3).value = deduct_sum
            ws.cell(row, 4).value = comment
    ws.cell(row+1, 1).value = "Additional Comments"

    wb.save(location + '/Rubric.xlsx')

def main():
    root = tk.Tk()
    root.withdraw()

    input_folder_path = filedialog.askdirectory()
    if input_folder_path == '':
        return

    outputwb = openpyxl.Workbook()
    outputws = outputwb['Sheet']

    outputws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 30
    outputws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 15
    outputws.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 60

    rubrics = {}
    for filename in os.listdir(input_folder_path):  # loop through all the files and folders
        if os.path.isdir(
                os.path.join(input_folder_path, filename)):  # check whether the current object is a folder or not
            if os.path.isfile(os.path.join(input_folder_path,filename, 'Rubric.xlsx')):
                rubrics[filename] = os.path.join(input_folder_path,filename, 'Rubric.xlsx')

    for index, name in enumerate(rubrics.keys()):
        rubricpath = rubrics[name]
        rubricwb = openpyxl.load_workbook(rubricpath)
        rubricws = rubricwb.worksheets[0]
        split_name = name.split('_')
        name = split_name[0] + ', ' + split_name[1]
        comment_string = ''
        total_points_achieved = 0
        for row in range(2, rubricws.max_row+1):
            points_deducted = int(rubricws.cell(row, 3).value) if rubricws.cell(row, 3).value != None else 0
            points_possible = int(rubricws.cell(row, 2).value) if rubricws.cell(row, 2).value != None else 0
            points_achieved = points_possible - points_deducted
            total_points_achieved += points_achieved
            comment_string += rubricws.cell(row, 1).value
            if points_possible != 0:
                comment_string += ' [' + str(points_achieved) + '/' + str(points_possible) + ']'
            if rubricws.cell(row, 4).value != None:
                comment = ':\n' + rubricws.cell(row, 4).value
                comment.replace('\n', '\n\t')
                comment_string += comment
            elif row < rubricws.max_row+1:
                comment_string += '\n'
        outputws.cell(index+1, 1).value = name
        outputws.cell(index+1, 2).value = total_points_achieved
        outputws.cell(index+1, 3).value = comment_string

    outputwb.save(os.path.join(input_folder_path, 'AllStudents.xlsx'))


if __name__ == "__main__":
    main()
    print("done")
