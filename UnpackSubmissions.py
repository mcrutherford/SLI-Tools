"""

@file: UnpackSubmissions.py
@author: Mark Rutherford [ mcr5801 ]
@created: 8/29/2018 9:06 PM
@description:
    Unpack student submissions to individual folders with personal xlsx grading files.

    Validated for Python 3.8, but should work for Python 3.x. Requires openpyxl, which can be installed with the command
    'pip install openpyxl'

"""
import openpyxl  # pip install openpyxl
import openpyxl.utils  # pip install openpyxl
import zipfile
import tkinter as tk
from tkinter import filedialog
import os
import re
import json


def create_individual_xl(rubric_data, location):
    """
    Create an individual rubric xlsx for the student.

    :param rubric_data: The loaded json data from the rubric
    :param location: The file in which to save the rubric
    """
    wb = openpyxl.Workbook()
    ws = wb['Sheet']

    # set the column widths to make the text more visible
    ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 35
    ws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 11
    ws.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 70

    # Name the columns
    col_names = ['Requirement', 'Points available', 'Deductions', 'Comments']
    for i in range(len(col_names)):
        ws.cell(1, i+1).value = col_names[i]

    # Insert the requirements
    row = 1
    for i, key in enumerate(rubric_data['requirements'].keys()):
        row = i+2
        ws.cell(row, 1).value = key
        ws.cell(row, 2).value = rubric_data['requirements'][key]["value"]

    # Add a slot for additional comments
    ws.cell(row+1, 1).value = "Additional Comments"

    # Save the file
    wb.save(os.path.join(location, 'Rubric.xlsx'))


def parse_json(file_location):
    """
    Load a json file into a dictionary.

    :param file_location: The file to load
    :return: A dictionary of the loaded data
    """
    file = open(file_location, 'r')
    json_data = json.loads(file.read())
    file.close()

    return json_data


def main():
    # Get the input and output file paths
    root = tk.Tk()
    root.withdraw()
    input_file_path = filedialog.askopenfilename(title='Select the zip of student assignments', filetypes=[('ZIP', '*.zip')])
    if input_file_path == '':
        return
    output_file_path = os.path.dirname(os.path.abspath(input_file_path))

    # Get the rubric
    rubric_path = filedialog.askopenfilename(title='Select the rubric json', filetypes=[('JSON', '*.json')])
    if rubric_path == '':
        return
    rubric_data = parse_json(rubric_path)

    # Extract the main zip file
    zip_ref = zipfile.ZipFile(input_file_path, 'r')
    zip_ref.extractall(output_file_path)
    zip_ref.close()

    # Extract each student's zip file if possible
    print('Extracting files...')
    for filename in os.listdir(output_file_path):
        if filename == os.path.basename(input_file_path) or filename[-4:] != '.zip':
            # these files are not student zip files, and won't be extracted
            continue
        try:
            # Identify the student's name from the mycourses file naming scheme
            pattern = re.compile("\d*-\d*\s-\s(.*),\s(.*)\s-\s(.*)\.zip")
            matches = pattern.match(filename)
            if not matches:
                raise zipfile.BadZipFile
            else:
                # Extract the student's zip file
                student_name = matches[1] + "_" + matches[2]
                zip_name = matches[3]
                zip_ref = zipfile.ZipFile(os.path.join(output_file_path, filename), 'r')
                zip_ref.extractall(os.path.join(output_file_path, student_name, zip_name))
                zip_ref.close()

                # Remove the student's extracted zip file
                os.remove(os.path.join(output_file_path, filename))

                create_individual_xl(rubric_data, os.path.join(output_file_path, student_name))
        except zipfile.BadZipFile:  # if a file wasn't able to be extracted, log the error
            print("> Failed to extract", filename)


if __name__ == "__main__":
    main()
    print("\ndone")
